#!/usr/bin/env python3
"""Exporta rules.sqlite a un .xlsx con una hoja por caso de produccion.

Las hojas por caso son las que se usan a diario: abres T1 y ves las reglas de
un maestro de rostro. La columna ORIGEN dice de donde salio cada clasificacion,
que es lo que permite no tener que creerle a nadie.

Uso:
    python3 export_xlsx.py --db rules.sqlite --out reglas-produccion.xlsx
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FUENTE = "Arial"
AZUL = "1F3864"      # cabeceras
GRIS = "F2F2F2"      # bandas
AMBAR = "FFF2CC"     # clasificacion gruesa, menos confiable
VERDE = "E2EFDA"     # corregida por auditoria externa

CASOS_ORDEN = ["T1", "T2", "T3", "T4", "T5", "PROD", "GRAF", "MULTI", "REF",
               "CLIP", "SHOT", "GUION", "MARCA", "QA", "ENTREGA", "NINGUNO"]

BORDE = Border(*[Side(style="thin", color="D0D0D0")] * 4)


def cabecera(ws, titulos, ancho):
    for i, t in enumerate(titulos, 1):
        c = ws.cell(row=1, column=i, value=t)
        c.font = Font(name=FUENTE, bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = BORDE
        ws.column_dimensions[get_column_letter(i)].width = ancho[i - 1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(titulos))}1"


def escribir(ws, filas, resaltar_origen=True):
    for r, fila in enumerate(filas, start=2):
        for c, val in enumerate(fila, start=1):
            cel = ws.cell(row=r, column=c, value=val)
            cel.font = Font(name=FUENTE, size=10)
            cel.alignment = Alignment(vertical="top")
            cel.border = BORDE
        if resaltar_origen:
            origen = fila[-1]
            if origen == "archivo":
                relleno = PatternFill("solid", fgColor=AMBAR)
            elif origen == "auditoria":
                relleno = PatternFill("solid", fgColor=VERDE)
            else:
                relleno = PatternFill("solid", fgColor=GRIS) if r % 2 == 0 else None
            if relleno:
                for c in range(1, len(fila) + 1):
                    ws.cell(row=r, column=c).fill = relleno


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default="rules.sqlite")
    ap.add_argument("--out", default="reglas-produccion.xlsx")
    a = ap.parse_args()

    con = sqlite3.connect(a.db)
    wb = Workbook()

    # ---------------------------------------------------------- RESUMEN
    ws = wb.active
    ws.title = "RESUMEN"
    ws["A1"] = "Base de reglas del pipeline de produccion visual"
    ws["A1"].font = Font(name=FUENTE, bold=True, size=14, color=AZUL)
    ws["A2"] = "Una hoja por caso. Abre T1 para un maestro de rostro, CLIP para un prompt de video."
    ws["A2"].font = Font(name=FUENTE, size=10, italic=True)

    ws["A4"] = "COMO LEER LA COLUMNA ORIGEN"
    ws["A4"].font = Font(name=FUENTE, bold=True, size=11)
    leyenda = [
        ("regla", "Clasificada una por una. Mas confiable.", GRIS),
        ("auditoria", "Corregida por un criterio externo. Manda sobre las otras dos.", VERDE),
        ("archivo", "Clasificacion gruesa por defecto del archivo. Revisar antes de confiar.", AMBAR),
    ]
    for i, (cod, desc, color) in enumerate(leyenda, start=5):
        ws.cell(row=i, column=1, value=cod).font = Font(name=FUENTE, bold=True, size=10)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=i, column=2, value=desc).font = Font(name=FUENTE, size=10)

    fila = 10
    ws.cell(row=fila, column=1, value="CASO").font = Font(name=FUENTE, bold=True, color="FFFFFF")
    ws.cell(row=fila, column=1).fill = PatternFill("solid", fgColor=AZUL)
    for j, t in enumerate(["REGLAS", "DESCRIPCION"], start=2):
        ws.cell(row=fila, column=j, value=t).font = Font(name=FUENTE, bold=True, color="FFFFFF")
        ws.cell(row=fila, column=j).fill = PatternFill("solid", fgColor=AZUL)

    primera = fila + 1
    for cod in CASOS_ORDEN:
        r = con.execute(
            "SELECT (SELECT descripcion FROM casos WHERE codigo=?),"
            " (SELECT COUNT(*) FROM regla_caso WHERE caso=?)", (cod, cod)).fetchone()
        if not r[1]:
            continue
        fila += 1
        ws.cell(row=fila, column=1, value=cod).font = Font(name=FUENTE, bold=True, size=10)
        ws.cell(row=fila, column=2, value=r[1]).font = Font(name=FUENTE, size=10)
        ws.cell(row=fila, column=3, value=r[0]).font = Font(name=FUENTE, size=10)
    ws.cell(row=fila + 1, column=1, value="Suma de asignaciones").font = Font(name=FUENTE, bold=True)
    ws.cell(row=fila + 1, column=2, value=f"=SUM(B{primera}:B{fila})").font = Font(name=FUENTE, bold=True)

    fila += 3
    tot = con.execute("SELECT COUNT(*) FROM reglas").fetchone()[0]
    aud = con.execute("SELECT COUNT(*) FROM auditorias").fetchone()[0]
    for etiqueta, valor in [
        ("Reglas unicas en la base", tot),
        ("Correcciones de auditoria registradas", aud),
        ("Skills de origen", con.execute("SELECT COUNT(*) FROM skills").fetchone()[0]),
        ("Archivos .md con sha256", con.execute("SELECT COUNT(*) FROM archivos").fetchone()[0]),
        ("Conflictos entre reglas registrados", con.execute("SELECT COUNT(*) FROM conflictos").fetchone()[0]),
    ]:
        ws.cell(row=fila, column=1, value=etiqueta).font = Font(name=FUENTE, size=10)
        ws.cell(row=fila, column=2, value=valor).font = Font(name=FUENTE, bold=True, size=10)
        fila += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 62

    # ------------------------------------------------- UNA HOJA POR CASO
    for cod in CASOS_ORDEN:
        filas = con.execute(
            """SELECT r.skill, r.archivo, r.linea, r.texto, rc.origen
               FROM regla_caso rc JOIN reglas r ON r.id = rc.regla_id
               WHERE rc.caso = ? ORDER BY r.skill, r.archivo, r.linea""", (cod,)).fetchall()
        if not filas:
            continue
        ws = wb.create_sheet(cod)
        cabecera(ws, ["SKILL", "ARCHIVO", "LINEA", "REGLA", "ORIGEN"],
                 [24, 38, 8, 105, 12])
        escribir(ws, filas)

    # ------------------------------------------------------ TODAS LAS REGLAS
    ws = wb.create_sheet("TODAS")
    cabecera(ws, ["ID", "SKILL", "ARCHIVO", "LINEA", "REGLA", "CASOS", "ORIGEN"],
             [15, 24, 38, 8, 95, 34, 12])
    escribir(ws, con.execute(
        """SELECT r.id, r.skill, r.archivo, r.linea, r.texto,
                  GROUP_CONCAT(rc.caso), MIN(rc.origen)
           FROM reglas r LEFT JOIN regla_caso rc ON rc.regla_id = r.id
           GROUP BY r.id ORDER BY r.skill, r.archivo, r.linea""").fetchall())

    # ------------------------------------------------------------ AUDITORIAS
    ws = wb.create_sheet("AUDITORIAS")
    cabecera(ws, ["ID", "ANTES", "DESPUES", "AUDITOR", "RAZON", "REGLA"],
             [15, 30, 30, 14, 60, 85])
    escribir(ws, con.execute(
        """SELECT a.regla_id, a.antes, a.despues, a.auditor, a.razon, g.texto
           FROM auditorias a JOIN reglas g ON g.id = a.regla_id ORDER BY a.rowid"""
        ).fetchall(), resaltar_origen=False)

    # ------------------------------------------------------------ CONFLICTOS
    filas = con.execute(
        """SELECT ra.texto, rb.texto, c.tipo, c.autoridad
           FROM conflictos c JOIN reglas ra ON ra.id = c.regla_a
           JOIN reglas rb ON rb.id = c.regla_b""").fetchall()
    if filas:
        ws = wb.create_sheet("CONFLICTOS")
        cabecera(ws, ["REGLA A (gana)", "REGLA B", "TIPO", "AUTORIDAD"], [70, 70, 16, 70])
        escribir(ws, filas, resaltar_origen=False)

    # --------------------------------------------------------------- FUENTES
    ws = wb.create_sheet("FUENTES")
    cabecera(ws, ["SKILL", "ARCHIVO", "REGLAS", "SHA256"], [26, 46, 10, 68])
    escribir(ws, con.execute(
        "SELECT skill, ruta, reglas, sha256 FROM archivos ORDER BY skill, ruta").fetchall(),
        resaltar_origen=False)

    Path(a.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(a.out)
    print(f"{a.out}: {len(wb.sheetnames)} hojas -> {', '.join(wb.sheetnames)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
